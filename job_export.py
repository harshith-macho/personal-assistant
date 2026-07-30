#!/usr/bin/env python3
"""
Job Export
- Copies each ATS-tailored resume into a human-facing folder, named
  CompanyName_RoleName_HarshithMittapally_Date_Resume.pdf
- Appends one row per job to an Excel log (openpyxl)
"""

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

MY_NAME = "HarshithMittapally"

OUTPUT_DIR   = Path("/Users/harshithreddy/Developer/MyAssist/JOBSRESUME _BOT")
RESUMES_DIR  = OUTPUT_DIR / "COM_ROLE_JOBS"
EXCEL_PATH   = OUTPUT_DIR / "Job_Applications_Log.xlsx"

HEADERS = ["Company Name", "Role Name", "Applied Date", "Resume File Name", "Ref ID", "Job ID"]

RECRUITER_EXCEL_PATH = OUTPUT_DIR / "Recruiter_Contacts_Log.xlsx"
RECRUITER_HEADERS = ["Name", "Company", "Medium", "Contact Link", "Role (Job Applied)", "Date Found", "Ref ID", "Job ID"]


def _sanitize(text: str, max_len: int = 40) -> str:
    """Strip to alphanumerics only so it's safe as a filename component."""
    cleaned = re.sub(r"[^A-Za-z0-9]+", "", text or "")
    return cleaned[:max_len] or "Unknown"


def _ensure_workbook():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    RESUMES_DIR.mkdir(parents=True, exist_ok=True)
    if not EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)


def _ensure_recruiter_workbook():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not RECRUITER_EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Recruiters"
        ws.append(RECRUITER_HEADERS)
        wb.save(RECRUITER_EXCEL_PATH)


def _get_ref_id(job_id: str) -> str:
    try:
        from job_tracker import ensure_ref_ids
        import sqlite3
        ensure_ref_ids()
        conn = sqlite3.connect(Path(__file__).parent / "applied_jobs.db")
        row = conn.execute("SELECT ref_id FROM jobs WHERE id=?", (job_id,)).fetchone()
        conn.close()
        return row[0] if row and row[0] else ""
    except Exception:
        return ""


def _wants_cv(job_id: str) -> bool:
    """True if the job's cached description explicitly asks for a 'CV' and not a 'resume'."""
    try:
        import sqlite3
        conn = sqlite3.connect(Path(__file__).parent / "applied_jobs.db")
        row = conn.execute("SELECT description FROM jobs WHERE id=?", (job_id,)).fetchone()
        conn.close()
        jd = (row[0] or "") if row else ""
        has_cv     = re.search(r"\bCVs?\b", jd) is not None
        has_resume = re.search(r"\br[eé]sum[eé]s?\b", jd, re.IGNORECASE) is not None
        return has_cv and not has_resume
    except Exception:
        return False


def readable_filename(job: dict, label: str = "Resume") -> str:
    """Human-readable resume filename, e.g. CompanyX_JobA_Resume.pdf — same
    convention as the archived copy in COM_ROLE_JOBS, for use anywhere a
    tailored resume is shown to the user (e.g. sent via Telegram)."""
    company = _sanitize(job.get("company", ""))
    role    = _sanitize(job.get("title", "").splitlines()[0])
    return f"{company}_{role}_{label}.pdf"


def export_resume(job: dict, tailored_pdf_path: str) -> str | None:
    """
    Copy the tailored PDF to RESUMES_DIR with the naming convention and
    append a row to the Excel log. Returns the saved path, or None on failure.
    """
    try:
        _ensure_workbook()

        company = _sanitize(job.get("company", ""))
        role    = _sanitize(job.get("title", "").splitlines()[0])
        date    = datetime.now().strftime("%Y-%m-%d")
        ref_id  = _get_ref_id(job.get("id", ""))
        label   = "CV" if _wants_cv(job.get("id", "")) else "Resume"

        filename = f"{company}_{role}_{MY_NAME}_{date}_{label}.pdf"
        dest = RESUMES_DIR / filename

        shutil.copy2(tailored_pdf_path, dest)

        wb = load_workbook(EXCEL_PATH)
        ws = wb["Applications"]
        job_id = job.get("id", "")
        new_row = [job.get("company", ""), job.get("title", "").splitlines()[0],
                   date, filename, ref_id, job_id]

        # Update in place if this job was already logged (e.g. resume regenerated), else append
        updated = False
        for row in ws.iter_rows(min_row=2):
            if row[5].value == job_id:
                for cell, value in zip(row, new_row):
                    cell.value = value
                updated = True
                break
        if not updated:
            ws.append(new_row)
        wb.save(EXCEL_PATH)

        return str(dest)
    except Exception as e:
        print(f"[job_export] error: {e}")
        return None


def export_recruiter(job: dict, recruiter: dict) -> bool:
    """
    Append a recruiter contact row: Name, Company, Medium, Contact Link,
    Role applied for, Date found, Ref ID. Skips rows already logged for the
    same profile link + job. Returns True on success.
    """
    try:
        _ensure_recruiter_workbook()

        name        = recruiter.get("name", "")
        profile_url = recruiter.get("profile_url", "")
        company     = job.get("company", "")
        role        = job.get("title", "").splitlines()[0]
        ref_id      = _get_ref_id(job.get("id", ""))
        date        = datetime.now().strftime("%Y-%m-%d")

        wb = load_workbook(RECRUITER_EXCEL_PATH)
        ws = wb["Recruiters"]

        # Dedup: same profile link already logged against this job
        job_id = job.get("id", "")
        key = (profile_url or name, job_id)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[3] or row[0], row[7]) == key:
                return True

        ws.append([name, company, "LinkedIn", profile_url, role, date, ref_id, job_id])
        wb.save(RECRUITER_EXCEL_PATH)
        return True
    except Exception as e:
        print(f"[job_export] recruiter export error: {e}")
        return False
